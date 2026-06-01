"""Generate formatted Excel reports from the Blue Mark SQLite database.

Produces reports in two formats:
  - v2026: Pre-pour / wythe / top check production checklist columns.
  - legacy: Post-pour QA tracking columns (2022-2025 data).

Usage:
    python scripts/generate_blue_mark_report.py --date 2026-01-27
    python scripts/generate_blue_mark_report.py --date 2026-01-27 --bed-type BMD
    python scripts/generate_blue_mark_report.py --job 12345 --format legacy
    python scripts/generate_blue_mark_report.py --active --format v2026
"""

import argparse
import logging
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Add parent to path so stresscon package is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from stresscon.blue_mark_db import BlueMarkDB

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions for each report format
# ---------------------------------------------------------------------------

V2026_COLUMNS: list[tuple[str, str, int]] = [
    # (Header text, dict key, column width)
    ("JOB NUMBER", "job_number", 12),
    ("PIECE NUMBER", "piece_number", 15),
    ("PRODUCTION PRE-POUR", "production_pre_pour", 20),
    ("QA PRE-POUR", "qa_pre_pour", 20),
    ("PRODUCTION WYTHE CHECK", "production_wythe_check", 20),
    ("QA WYTHE CHECK", "qa_wythe_check", 20),
    ("PRODUCTION TOP CHECK", "production_top_check", 20),
]

LEGACY_COLUMNS: list[tuple[str, str, int]] = [
    ("Pour Date", "pour_date", 14),
    ("Job #", "job_number", 12),
    ("Piece #", "piece_number", 15),
    ("Piece Blue Marked", "blue_marked", 18),
    ("Patching/Cleaning", "patching_cleaning", 18),
    ("Post Pour Weld-On Required", "weld_on_required", 26),
    ("NCR Response", "ncr_response", 14),
    ("NCR Repair", "ncr_repair", 14),
    ("Comments", "comments", 30),
]

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Checkpoint keys that store boolean-like integers (1 / 0)
CHECKBOX_KEYS = {
    "blue_marked",
    "patching_cleaning",
    "weld_on_required",
    "ncr_response",
    "ncr_repair",
    "production_pre_pour",
    "qa_pre_pour",
    "production_wythe_check",
    "qa_wythe_check",
    "production_top_check",
}


# ---------------------------------------------------------------------------
# Excel generation helpers
# ---------------------------------------------------------------------------

def _format_cell_value(key: str, raw_value: Any) -> Any:
    """Convert a raw database value into a display value for the spreadsheet."""
    if key in CHECKBOX_KEYS:
        return "X" if raw_value else ""
    if raw_value is None:
        return ""
    return raw_value


def _write_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[tuple[str, str, int]],
    pieces: list[dict[str, Any]],
) -> None:
    """Write a single worksheet with formatted headers and data rows."""
    ws = wb.create_sheet(title=sheet_name)

    # -- Header row ----------------------------------------------------------
    for col_idx, (header, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- Data rows -----------------------------------------------------------
    for row_idx, piece in enumerate(pieces, start=2):
        for col_idx, (_header, key, _width) in enumerate(columns, start=1):
            value = _format_cell_value(key, piece.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            # Center checkbox / "X" columns
            if key in CHECKBOX_KEYS:
                cell.alignment = CENTER_ALIGN

    # -- Auto-filter on headers ----------------------------------------------
    last_col = get_column_letter(len(columns))
    last_row = max(len(pieces) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _detect_format(pieces: list[dict[str, Any]]) -> str:
    """Determine the report format from the data's schema_version field.

    Returns 'v2026' if any piece uses that schema; otherwise 'legacy'.
    """
    for piece in pieces:
        if piece.get("schema_version") == "v2026":
            return "v2026"
    return "legacy"


def _build_filename(
    label: str,
    report_date: str | None,
    bed_type: str | None,
) -> str:
    """Build the output filename (without directory)."""
    if bed_type:
        return f"{bed_type}_{report_date or label}.xlsx"
    return f"Blue_Mark_Report_{report_date or label}.xlsx"


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def generate_report(
    pieces: list[dict[str, Any]],
    *,
    report_format: str,
    sheet_name: str,
    output_dir: Path,
    report_date: str | None = None,
    bed_type: str | None = None,
) -> Path:
    """Create a single Excel workbook from the given piece records.

    Args:
        pieces: List of piece dicts from BlueMarkDB.
        report_format: 'v2026' or 'legacy'.
        sheet_name: Name of the worksheet.
        output_dir: Directory in which to save the file.
        report_date: ISO date string used in the filename.
        bed_type: Optional bed type label used in the filename.

    Returns:
        Path to the written Excel file.
    """
    columns = V2026_COLUMNS if report_format == "v2026" else LEGACY_COLUMNS
    wb = Workbook()

    # Remove the default empty sheet created by openpyxl
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    _write_sheet(wb, sheet_name, columns, pieces)

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = _build_filename("Active_Pieces" if not report_date else report_date, report_date, bed_type)
    filepath = output_dir / filename
    wb.save(str(filepath))
    return filepath


def generate_reports_by_bed_type(
    pieces: list[dict[str, Any]],
    *,
    report_format: str,
    sheet_name: str,
    output_dir: Path,
    report_date: str | None = None,
) -> list[Path]:
    """Split pieces by bed_type and generate one file per group.

    Args:
        pieces: List of piece dicts from BlueMarkDB.
        report_format: 'v2026' or 'legacy'.
        sheet_name: Name of the worksheet.
        output_dir: Directory in which to save the files.
        report_date: ISO date string used in the filename.

    Returns:
        List of paths to the written Excel files.
    """
    groups: dict[str, list[dict[str, Any]]] = {}
    for piece in pieces:
        bed = piece.get("bed_type") or "UNKNOWN"
        groups.setdefault(bed, []).append(piece)

    paths: list[Path] = []
    for bed_type, bed_pieces in sorted(groups.items()):
        path = generate_report(
            bed_pieces,
            report_format=report_format,
            sheet_name=sheet_name,
            output_dir=output_dir,
            report_date=report_date,
            bed_type=bed_type,
        )
        paths.append(path)
    return paths


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate Blue Mark Excel reports from the SQLite database."
    )
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="Report date in YYYY-MM-DD format (default: today).",
    )
    parser.add_argument(
        "--job",
        type=int,
        default=None,
        help="Filter to a specific job number.",
    )
    parser.add_argument(
        "--bed-type",
        type=str,
        default=None,
        help="Filter to a specific bed type (e.g. BMD, S4, HICAP). "
             "When set, one file per bed type is generated.",
    )
    parser.add_argument(
        "--active",
        action="store_true",
        default=False,
        help="Show only pieces that haven't completed all checkpoints.",
    )
    parser.add_argument(
        "--format",
        dest="report_format",
        type=str,
        choices=["v2026", "legacy"],
        default=None,
        help="Report format: 'v2026' or 'legacy'. Default: auto-detect from data.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=".",
        help="Directory in which to save Excel files (default: current directory).",
    )
    parser.add_argument(
        "--db-path",
        type=str,
        default="data/blue_mark.db",
        help="Path to the Blue Mark SQLite database (default: data/blue_mark.db).",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    """Entry point for the Blue Mark report generator."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    args = parse_args(argv)

    db_path = Path(args.db_path)
    output_dir = Path(args.output_dir)

    # Determine the target date (default to today)
    if args.date:
        try:
            report_date = date.fromisoformat(args.date)
        except ValueError:
            logger.error("Invalid date format: %s. Use YYYY-MM-DD.", args.date)
            sys.exit(1)
    else:
        report_date = date.today()

    # -----------------------------------------------------------------------
    # Query the database
    # -----------------------------------------------------------------------
    db = BlueMarkDB(db_path=db_path)
    db.connect()

    try:
        if args.active:
            # For --active, the format hint selects the schema_version filter
            schema_hint = args.report_format  # may be None (query both)
            pieces = db.get_active_pieces(schema_version=schema_hint)
            logger.info("Fetched %d active piece(s).", len(pieces))
        elif args.job is not None:
            pieces = db.get_pieces_by_job(args.job)
            logger.info(
                "Fetched %d piece(s) for job %d.", len(pieces), args.job
            )
        else:
            pieces = db.get_pieces_by_date(report_date)
            logger.info(
                "Fetched %d piece(s) for date %s.", len(pieces), report_date.isoformat()
            )
    finally:
        db.close()

    if not pieces:
        logger.warning("No data found for the given query. No report generated.")
        sys.exit(0)

    # -----------------------------------------------------------------------
    # Apply optional bed-type filter
    # -----------------------------------------------------------------------
    if args.bed_type and not args.active:
        pieces = [p for p in pieces if p.get("bed_type") == args.bed_type]
        if not pieces:
            logger.warning(
                "No pieces matched bed type '%s'. No report generated.",
                args.bed_type,
            )
            sys.exit(0)

    # -----------------------------------------------------------------------
    # Determine report format
    # -----------------------------------------------------------------------
    report_format = args.report_format or _detect_format(pieces)
    logger.info("Using report format: %s", report_format)

    # -----------------------------------------------------------------------
    # Build the sheet name
    # -----------------------------------------------------------------------
    if args.active:
        sheet_name = "Active Pieces"
        report_date_str: str | None = None
    else:
        sheet_name = report_date.isoformat()
        report_date_str = report_date.isoformat()

    # -----------------------------------------------------------------------
    # Generate files
    # -----------------------------------------------------------------------
    if args.bed_type:
        # Single bed type requested -- one file with bed_type in filename
        path = generate_report(
            pieces,
            report_format=report_format,
            sheet_name=sheet_name,
            output_dir=output_dir,
            report_date=report_date_str,
            bed_type=args.bed_type,
        )
        print(f"Report saved: {path.resolve()}")
    else:
        # Check if data spans multiple bed types; if so, split into files
        bed_types_in_data = {p.get("bed_type") for p in pieces if p.get("bed_type")}
        if len(bed_types_in_data) > 1:
            paths = generate_reports_by_bed_type(
                pieces,
                report_format=report_format,
                sheet_name=sheet_name,
                output_dir=output_dir,
                report_date=report_date_str,
            )
            for p in paths:
                print(f"Report saved: {p.resolve()}")
        else:
            path = generate_report(
                pieces,
                report_format=report_format,
                sheet_name=sheet_name,
                output_dir=output_dir,
                report_date=report_date_str,
            )
            print(f"Report saved: {path.resolve()}")


if __name__ == "__main__":
    main()
